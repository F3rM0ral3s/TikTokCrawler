from selenium.webdriver.common.by import By
from openpyxl import Workbook
from openpyxl import load_workbook
from time import sleep
import winsound
import re
from datetime import datetime,timedelta


def createExcel():
    excel = Workbook()
    hoja = excel.active

    # Headers de columna
    headers = ["Descripción","URL","Fecha Publicación","Comentario","Likes","Fecha comment"]

    # Imprimimos en el excel
    for col,head in enumerate(headers):
         hoja.cell(row=1,column=col+1,value=head)

    # Fecha de analisis
    hoja.cell(row=1,column=8,value='Fecha de analisis:')
    hoja.cell(row=1,column=9,value=datetime.now())

    excel.save("videos_comments.xlsx")

def getFecha(fecha_str):
    try:
        if 'Hace' in fecha_str:
            num = int(re.findall('\d+',fecha_str)[0])
            if 'segundo' in fecha_str:
                return datetime.now()
            elif 'min' in fecha_str:
                return datetime.now()
            elif ('hora' in fecha_str) or ('h' in fecha_str):
                return datetime.now()
            elif 'día' in fecha_str:
                return datetime.now()-timedelta(days=num)
            elif 'semana' in fecha_str:
                return datetime.now()-timedelta(weeks=num)
            elif 'mes' in fecha_str:
                return datetime.now()-timedelta(months=num)
            else:
                return fecha_str

        else:
            nums = re.split('-',fecha_str)
            if(len(nums)==2):
                return datetime(2024,int(nums[0]),int(nums[1]))
            else:
                return datetime(int(nums[2]),int(nums[0]),int(nums[1]))
    except Exception as e:
        print(e)
        print(' [ ERROR EN FECHA ]')
        return fecha_str

def getComments(driver,url,desc,other_info):
    #Inicializamos lista de comentarios
    coms = []
   
    # Accedemos al excel
    excel = load_workbook("videos_comments.xlsx")
    hoja = excel.active
    curr = hoja.max_row + 1

    try:
        # Escribimos la descripcion, la url y tratamos la fecha
        hoja.cell(row=curr,column=1,value=desc)
        hoja.cell(row=curr,column=2,value=url)
        hoja.cell(row=curr,column=3,value=getFecha(re.split('\n',other_info)[2]))
        
        # Variables auxiliares
        start = 0

        # Encuentra todos los comentarios visibles
        coms = driver.find_elements(By.CSS_SELECTOR,'div.css-ulyotp-DivCommentContentContainer.e1g2efjf0')
        while(True):
            # Acotamos la lista de comentarios
            aux = coms
            coms = coms[start:]
    
            # Escribimos los comments
            for com in coms:
                driver.execute_script("arguments[0].scrollIntoView(true);", com)
                # Tenemos que hacerlo ligado para estar seguro que la cantidad
                # de likes y el texto corresponden al mismo comentario

                # Obtenemos el texto
                texto = com.find_element(By.CSS_SELECTOR,'div.css-1mf23fd-DivContentContainer.e1g2efjf1')
                texto = texto.find_element(By.CSS_SELECTOR,'p.css-xm2h10-PCommentText.e1g2efjf6')
                
                # Obtenemos numero de likes
                likes = com.find_element(By.CSS_SELECTOR,'div.css-1swe2yf-DivActionContainer.esns4rh0')
                likes = likes.find_element(By.CSS_SELECTOR,'div.css-114tc9h-DivLikeWrapper.ezxoskx0')
                likes = likes.get_attribute('aria-label')

                # Obtenemos la fecha del comment
                fecha_com = com.find_element(By.CSS_SELECTOR,'div.css-1mf23fd-DivContentContainer.e1g2efjf1')
                fecha_com = fecha_com.find_element(By.CSS_SELECTOR,'span.css-4tru0g-SpanCreatedTime.e1g2efjf8')
                
                # Escribimos el texto, likes y fecha de comment
                hoja.cell(row=curr,column=4,value=texto.text)
                hoja.cell(row=curr,column=5,value=re.findall('\d+',likes)[0])
                hoja.cell(row=curr,column=6,value=getFecha(fecha_com.text))
                curr+=1
                start+=1

                #Buscamos si apareció un captcha
                try:
                    driver.find_element(By.CSS_SELECTOR,"div.captcha_verify_container.style__CaptchaWrapper-sc-1gpeoge-0.zGYIR") 
        
                    # Esperamos respuesta del usuario
                    winsound.Beep(440,2000)
                    response = input('[COMMENTS] Resolviste el captcha [y/n]: ')
                    if(response not in ['y','Y']):
                        driver.quit()

                except Exception as e:
                    pass

            # Buscamos nuevos comments
            sleep(2)
            coms = driver.find_elements(By.CSS_SELECTOR,'div.css-ulyotp-DivCommentContentContainer.e1g2efjf0')
            if(len(coms)==len(aux) or start >=500):
                break
            
        # Guardamos
        excel.save("videos_comments.xlsx")   

    except Exception as e:
        print('[ Error en comments ]')
        print('[ Se salvaron %d comentarios ]' % len(coms))
        print(e)
        for com in coms:
                # Tenemos que hacerlo ligado para estar seguro que la cantidad
                # de likes y el texto corresponden al mismo comentario

                # Obtenemos el texto
                texto = com.find_element(By.CSS_SELECTOR,'div.css-1mf23fd-DivContentContainer.e1g2efjf1')
                texto = texto.find_element(By.CSS_SELECTOR,'p.css-xm2h10-PCommentText.e1g2efjf6')
                
                # Obtenemos numero de likes
                likes = com.find_element(By.CSS_SELECTOR,'div.css-1swe2yf-DivActionContainer.esns4rh0')
                likes = likes.find_element(By.CSS_SELECTOR,'div.css-114tc9h-DivLikeWrapper.ezxoskx0')
                likes = likes.get_attribute('aria-label')

                # Obtenemos la fecha del comment
                fecha_com = com.find_element(By.CSS_SELECTOR,'div.css-1mf23fd-DivContentContainer.e1g2efjf1')
                fecha_com = fecha_com.find_element(By.CSS_SELECTOR,'span.css-4tru0g-SpanCreatedTime.e1g2efjf8')
                
                # Escribimos el texto, likes y fecha de comment
                hoja.cell(row=curr,column=4,value=texto.text)
                hoja.cell(row=curr,column=5,value=re.findall('\d+',likes)[0])
                hoja.cell(row=curr,column=6,value=getFecha(fecha_com.text))
                curr+=1
                start+=1
        excel.save("videos_comments.xlsx")          